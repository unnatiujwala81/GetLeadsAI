import customtkinter as ctk
import pandas as pd
import webbrowser

from tkinter import ttk
from tkinter import filedialog


class Results(ctk.CTkFrame):

    def __init__(self, parent):
        super().__init__(parent)

        # Stores the real URLs for each table row
        self.row_links = {}

        self.build_ui()

    def build_ui(self):

        # =====================================
        # TITLE
        # =====================================

        title = ctk.CTkLabel(
            self,
            text="Results",
            font=("Segoe UI", 28, "bold")
        )

        title.pack(
            anchor="w",
            padx=20,
            pady=(20, 10)
        )

        # =====================================
        # TOOLBAR
        # =====================================

        toolbar = ctk.CTkFrame(self)
        toolbar.pack(
            fill="x",
            padx=20,
            pady=(0, 10)
        )

        self.total_label = ctk.CTkLabel(
            toolbar,
            text="Total Leads : 0",
            font=("Segoe UI", 15, "bold")
        )

        self.total_label.pack(
            side="left",
            padx=15
        )

        clear_btn = ctk.CTkButton(
            toolbar,
            text="🗑 Clear",
            width=100,
            command=self.clear_table,
            fg_color="#D32F2F",
            hover_color="#B71C1C"
        )

        clear_btn.pack(
            side="right",
            padx=5
        )

        export_csv_btn = ctk.CTkButton(
            toolbar,
            text="📄 Export CSV",
            width=120,
            command=self.export_csv
        )

        export_csv_btn.pack(
            side="right",
            padx=5
        )

        export_excel_btn = ctk.CTkButton(
            toolbar,
            text="📗 Export Excel",
            width=140,
            command=self.export_excel
        )

        export_excel_btn.pack(
            side="right",
            padx=5
        )

        # =====================================
        # TABLE FRAME
        # =====================================

        table_frame = ctk.CTkFrame(self)

        table_frame.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=10
        )

        style = ttk.Style()

        try:
            style.theme_use("clam")
        except:
            pass

        style.configure(
            "Treeview",
            background="#2B2B2B",
            foreground="white",
            fieldbackground="#2B2B2B",
            rowheight=30,
            font=("Segoe UI", 10)
        )

        style.configure(
            "Treeview.Heading",
            background="#1F6AA5",
            foreground="white",
            font=("Segoe UI", 10, "bold")
        )

        columns = (
            "Business Name",
            "Category",
            "Address",
            "Location",
            "Phone",
            "Email",
            "Website",
            "Facebook",
            "Instagram",
            "LinkedIn",
            "YouTube",
            "Google Rating",
            "Reviews",
            "Business Status",
            "Google Maps URL",
            "Notes"
        )

        self.table = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings"
        )

        widths = {
            "Business Name":300,
            "Category":240,
            "Address":300,
            "Location":250,
            "Phone":200,
            "Email":300,
            "Website":120,
            "Facebook":120,
            "Instagram":120,
            "LinkedIn":120,
            "YouTube":120,
            "Google Rating":40,
            "Reviews":80,
            "Business Status":250,
            "Google Maps URL":250,
            "Notes":300
        }

        for col in columns:

            self.table.heading(col, text=col)

            self.table.column(
                col,
                width=widths[col],
                anchor="center",
                stretch=True
            )

        y_scroll = ttk.Scrollbar(
            table_frame,
            orient="vertical",
            command=self.table.yview
        )

        x_scroll = ttk.Scrollbar(
            table_frame,
            orient="horizontal",
            command=self.table.xview
        )

        self.table.configure(
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )

        self.table.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        y_scroll.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        x_scroll.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Double-click to open links
        self.table.bind(
            "<Double-1>",
            self.open_link
      )
    # =====================================
    # ADD LEAD
    # =====================================

    def add_lead(self, lead):

        # Display short labels instead of long URLs
        values=(
            lead.get("Business Name", ""),
            lead.get("Category", ""),
            lead.get("Address", ""),
            lead.get("Location", ""),
            lead.get("Phone", ""),
            lead.get("Email", ""),
            
            "🌐 Visit" if lead.get("Website", "") else "",
            "📘 Facebook" if lead.get("Facebook", "") else "",
            "📷 Instagram" if lead.get("Instagram", "") else "",
            "💼 LinkedIn" if lead.get("LinkedIn", "") else "",
            "▶ YouTube" if lead.get("YouTube", "") else "",

            lead.get("Google Rating", ""),
            lead.get("Reviews", ""),
            lead.get("Business Status", ""),
            
            "📍 View Map"
            if lead.get("Google Maps URL", "")
            else "",

            lead.get("Notes", "")
        )

        # Insert the row and get its unique ID
        item_id = self.table.insert(
            "",
            "end",
            values=values
        )

        # Store the real URLs separately
        self.row_links[item_id] = {
            "Website": lead.get("Website", ""),
            "Facebook": lead.get("Facebook", ""),
            "Instagram": lead.get("Instagram", ""),
            "LinkedIn": lead.get("LinkedIn", ""),
            "YouTube": lead.get("YouTube", ""),
            "Google Maps URL": lead.get(
                "Google Maps URL",
                ""
           )
        }

        total = len(self.table.get_children())

        self.total_label.configure(
            text=f"Total Leads : {total}"
        )    
        
        self.total_label.configure(
            text=f"Total Leads : {len(self.table.get_children())}"
        )

    def open_link(self, event):

        # Find the clicked row
        item_id = self.table.identify_row(event.y)

        if not item_id:
            return

        # Find the clicked column
        column = self.table.identify_column(event.x)

        if not column:
            return

        column_index = int(
            column.replace("#", "")
        ) - 1

        # Column positions in the table
        link_columns = {
            6: "Website",
            7: "Facebook",
            8: "Instagram",
            9: "LinkedIn",
            10: "YouTube",
            14: "Google Maps URL"
        }

        # Do nothing when another column is clicked
        if column_index not in link_columns:
            return

        link_name = link_columns[column_index]

        # Get the real URL stored for this row
        url = self.row_links.get(
           item_id,
            {}
        ).get(
            link_name,
            ""
        )

        if url:

            # Add https:// if missing
            if not url.startswith(
                ("http://", "https://")
            ):
                url = "https://" + url

            webbrowser.open_new_tab(url)

        item = self.table.identify_row(event.y)

        if not item:
            return

        column = self.table.identify_column(event.x)

        values = self.table.item(item)["values"]

        col = int(column.replace("#", "")) - 1

        # Website
        if col == 6:
            url = values[6]

        # Facebook
        elif col == 7:
            url = values[7]

        # Instagram
        elif col == 8:
            url = values[8]

        # LinkedIn
        elif col == 9:
            url = values[9]

        # YouTube
        elif col == 10:
            url = values[10]

        # Google Maps
        elif col == 14:
            url = values[14]

        else:
            return

        if url:
            webbrowser.open(url)

    # =====================================
    # LOAD DATAFRAME
    # =====================================

    def load_dataframe(self, df):

        self.clear_table()

        for _, row in df.iterrows():

            self.add_lead({
                "Business Name": row.get("Business Name", ""),
                "Category": row.get("Category", ""),
                "Address": row.get("Address", ""),
                "Location": row.get("Location", ""),
                "Phone": row.get("Phone", ""),
                "Email": row.get("Email", ""),
                "Website": row.get("Website", ""),
                "Facebook": row.get("Facebook", ""),
                "Instagram": row.get("Instagram", ""),
                "LinkedIn": row.get("LinkedIn", ""),
                "YouTube": row.get("YouTube", ""),
                "Google Rating": row.get("Google Rating", ""),
                "Reviews": row.get("Reviews", ""),
                "Business Status": row.get("Business Status", ""),
                "Google Maps URL": row.get("Google Maps URL", ""),
                "Notes": row.get("Notes", "")
            })

    # =====================================
    # CLEAR TABLE
    # =====================================

    def clear_table(self):

        for row in self.table.get_children():
            self.table.delete(row)

        # Remove stored URLs too
        self.row_links.clear()    

        self.total_label.configure(
            text="Total Leads : 0"
        )

    # =====================================
    # EXPORT CSV
    # =====================================

    def export_csv(self):

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")]
        )

        if not filename:
            return

        rows = []

        for item in self.table.get_children():
            
            values = list(self.table.item(item)["values"])
            links = self.row_links.get(item, {})

            values[6] = links.get("Website", "")
            values[7] = links.get("Facebook", "")
            values[8] = links.get("Instagram", "")
            values[9] = links.get("LinkedIn", "")
            values[10] = links.get("YouTube", "")
            values[14] = links.get("Google Maps URL", "")

            rows.append(values)

        df = pd.DataFrame(rows, columns=self.table["columns"])
        df.to_csv(filename, index=False)

    # =====================================
    # EXPORT EXCEL
    # =====================================

    from openpyxl.styles import Font, PatternFill, Alignment

    def export_excel(self):

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if not filename:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.freeze_panes = "A2"

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 25
        ws.column_dimensions["O"].width = 20
        ws.column_dimensions["P"].width = 30
    
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

        row_fill = PatternFill(
        fill_type="solid",
        start_color="F8F9FA",
        end_color="F8F9FA"
        )

        header_font = Font(bold=True)

        columns = self.table["columns"]

        link_text = {
            7: "🌐 Visit",
            8: "📘 Facebook",
            9: "📷 Instagram",
            10: "💼 LinkedIn",
            11: "▶ YouTube",
            15: "📍 View Map"
        }

        # Headers
        header_font = Font(bold=True)

        for col_num, heading in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = heading
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Data
        for row_num, item in enumerate(self.table.get_children(), start=2):

            values = list(self.table.item(item)["values"])
            links = self.row_links.get(item, {})

            values[6] = links.get("Website", "")
            values[7] = links.get("Facebook", "")
            values[8] = links.get("Instagram", "")
            values[9] = links.get("LinkedIn", "")
            values[10] = links.get("YouTube", "")
            values[14] = links.get("Google Maps URL", "")

            for col_num, value in enumerate(values, start=1):

                cell = ws.cell(row=row_num, column=col_num)

                if (
                    isinstance(value, str)
                    and value.startswith(("http://", "https://"))
                    and col_num in link_text
                ):
                    cell.value = link_text[col_num]
                    cell.hyperlink = value
                    cell.font = Font(
                        bold=True,
                        color="0563C1",
                        underline="single"
                    )
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
                else:
                    cell.value = value

                    if row_num % 2 == 0:
                        cell.fill = row_fill

            # Add filter to header
            ws.auto_filter.ref = ws.dimensions

            # Save Excel file
            wb.save(filename)
